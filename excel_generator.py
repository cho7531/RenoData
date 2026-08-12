import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from extractor import COLUMNS, lookup_postal_codes

ADDRESS_COLUMN = "소유자 주소"
POSTAL_COLUMN = "우편번호"


def build_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "등기부등본 추출결과"

    ws.append(COLUMNS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        max_len = len(col_name)
        for row in rows:
            value = str(row.get(col_name, ""))
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def fill_postal_codes(file_stream):
    """Reads an uploaded .xlsx, looks up postal codes for each row's owner
    address, fills in any blank 우편번호 cells, and returns the updated
    workbook as a BytesIO buffer."""
    wb = load_workbook(file_stream)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    if ADDRESS_COLUMN not in header:
        raise ValueError(f"'{ADDRESS_COLUMN}' 열을 찾을 수 없습니다. 추출 결과 엑셀 형식이 맞는지 확인해주세요.")
    if POSTAL_COLUMN not in header:
        raise ValueError(f"'{POSTAL_COLUMN}' 열을 찾을 수 없습니다. 추출 결과 엑셀 형식이 맞는지 확인해주세요.")

    address_col = header.index(ADDRESS_COLUMN) + 1
    postal_col = header.index(POSTAL_COLUMN) + 1

    data_rows = list(ws.iter_rows(min_row=2))
    addresses = [
        row[address_col - 1].value
        for row in data_rows
        if row[address_col - 1].value and not row[postal_col - 1].value
    ]
    postal_codes = lookup_postal_codes(addresses)

    for row in data_rows:
        addr_cell = row[address_col - 1]
        postal_cell = row[postal_col - 1]
        if addr_cell.value and not postal_cell.value:
            postal_cell.value = postal_codes.get(str(addr_cell.value).strip(), "")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
